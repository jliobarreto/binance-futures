# notifier.py
import os
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
from utils.path import XLSX_PATH

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

# XLSX_PATH ya contiene la ruta completa del archivo, por lo que solo
# es necesario convertirlo a cadena para usarlo con openpyxl.
SIGNAL_FILE = str(XLSX_PATH)


def enviar_telegram(texto: str, buttons: list = None):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": texto,
        "parse_mode": "Markdown"
    }

    if buttons:
        keyboard = [[{"text": b, "callback_data": b}] for b in buttons]
        payload["reply_markup"] = {"inline_keyboard": keyboard}

    try:
        response = requests.post(url, json=payload, timeout=10)
        return response.json()
    except Exception as e:
        print(f"❌ Error enviando a Telegram: {e}")
        return None


wb = Workbook()
        ws = wb.active
        ws.append([
            "Fecha", "Criptomoneda", "Tipo", "Entrada", "TP", "SL", "RSI", "MACD",
            "Vitalidad", "Grids", "Score", "Decisión"
        ])
    else:
        wb = load_workbook(SIGNAL_FILE)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        op.get("Criptomoneda"),
        op.get("Señal"),
        op.get("Precio"),
        op.get("TP"),
        op.get("SL"),
        op.get("RSI"),
        op.get("MACD"),
        op.get("Vitalidad"),
        op.get("Grids"),
        op.get("Score"),
        decision
    ])
    wb.save(SIGNAL_FILE)
    print(f"✅ Operación guardada como '{decision}' en el archivo Excel")


def manejar_callback(callback_data: str, symbol: str, memoria: dict) -> None:
    """Procesa la respuesta del usuario desde Telegram.

    Parameters
    ----------
    callback_data : str
        Texto enviado por Telegram en el callback. Se espera que tenga la forma
        "DECISION|SYMBOL".
    symbol : str
        Símbolo de la operación asociada al mensaje.
    memoria : dict
        Estructura en memoria que almacena las operaciones enviadas.
    """

    # La decisión corresponde a la primera parte del callback
    decision = callback_data.split("|")[0]

    # Recuperar la operación de la memoria de operaciones enviadas
    operacion = memoria.pop(symbol, None)
    if not operacion:
        print(f"⚠️ Operación para {symbol} no encontrada en memoria")
        return

    guardar_operacion(operacion, decision)
    enviar_telegram(f"Respuesta recibida para {symbol}: {decision}")