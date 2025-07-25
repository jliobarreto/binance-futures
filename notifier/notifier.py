# notifier.py
import os
import logging
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
from utils.path import XLSX_PATH
from config import TELEGRAM_TOKEN, TELEGRAM_CHAT_ID
from logic.risk_manager import registrar_resultado

# XLSX_PATH ya contiene la ruta completa del archivo, por lo que solo
# es necesario convertirlo a cadena para usarlo con openpyxl.
SIGNAL_FILE = str(XLSX_PATH)


def enviar_telegram(texto: str, buttons: list = None):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": texto,
        "parse_mode": "Markdown"
    }

    if buttons:
        keyboard = [[{"text": b, "callback_data": b}] for b in buttons]
        payload["reply_markup"] = {"inline_keyboard": keyboard}

    try:
        response = requests.post(url, json=payload, timeout=10)
        return response.json()
    except Exception as e:
        logging.error(f"Error enviando a Telegram: {e}")
        return None


def guardar_operacion(op: dict, decision: str) -> None:
    """Guarda la operación y la decisión del usuario en un archivo Excel."""
    if not os.path.exists(SIGNAL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Fecha", "Criptomoneda", "Tipo", "Entrada", "TP", "SL", "RSI",
            "MACD", "Vitalidad", "Grids", "Score", "Decisión"
        ])
    else:
        wb = load_workbook(SIGNAL_FILE)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        op.get("Criptomoneda"),
        op.get("Señal"),
        op.get("Precio"),
        op.get("TP"),
        op.get("SL"),
        op.get("RSI"),
        op.get("MACD"),
        op.get("Vitalidad"),
        op.get("Grids"),
        op.get("Score"),
        decision
    ])
    wb.save(SIGNAL_FILE)
    logging.info(f"Operación guardada como '{decision}' en el archivo Excel")


def manejar_callback(callback_data: str, symbol: str, memoria: dict) -> None:
    """Procesa la respuesta del usuario desde Telegram."""

    partes = callback_data.split("|")
    decision = partes[0]

    # Permite registrar el resultado final de una operación con el formato
    # "RESULTADO|SYMBOL|PNL" enviado desde Telegram.
    if decision.lower() == "resultado" and len(partes) >= 3:
        try:
            pnl = float(partes[2])
        except ValueError:
            logging.error(f"PNL inválido en callback: {callback_data}")
            return
        registrar_resultado(pnl)
        enviar_telegram(
            f"Resultado para {symbol} registrado: {pnl:.2f} USDT"
        )
        return

    operacion = memoria.get(symbol)

    if not operacion:
        logging.error(f"Operación para {symbol} no encontrada en memoria")
        return

    # Registrar la decisión en el archivo Excel
    guardar_operacion(operacion, decision)

    # Enviar confirmación de registro al usuario
    enviar_telegram(f"Operación para {symbol} guardada como '{decision}'")

    # Eliminar la operación de la memoria para evitar duplicados
    memoria.pop(symbol, None)
